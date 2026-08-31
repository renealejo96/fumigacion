import io
import datetime
import pandas as pd
from app.extensions import db
from app.shared.models import (
    Rotation, Requisition, RequisitionItem, 
    FumigationOrder, FumigationOrderProductSummary, 
    AdditionalApplication, Product
)
from app.modules.fumigacion.services.calculation_engine import CalculationEngine
from app.shared.utils import is_integer_unit, round_product_amount

class RequisitionService:

    @classmethod
    def generate_or_update_forecast(cls, rotation_id: int, title: str = None, cached_crop_records=None) -> Requisition:
        """
        Generates or updates the 15-day advance purchase requisition from the rotation plan.
        """
        rotation = Rotation.query.get_or_404(rotation_id)
        
        product_totals = {}
        total_liters_all = 0.0

        for r in rotation.rounds:
            calc = CalculationEngine.calculate_round(r, cached_crop_records=cached_crop_records)
            total_liters_all += calc['totals']['total_liters']

            for ps in calc['product_summaries']:
                key = ps['product_code']
                if key not in product_totals:
                    product_totals[key] = {
                        'product_id': ps['product_id'],
                        'product_code': ps['product_code'],
                        'commercial_name': ps['commercial_name'],
                        'dose': ps['dose'],
                        'dose_unit': ps['dose_unit'],
                        'quantity_forecast': 0.0,
                        'pest': ps['pest']
                    }
                product_totals[key]['quantity_forecast'] += ps['total_required_quantity']

        # Find or create requisition
        req = Requisition.query.filter_by(rotation_id=rotation.id).first()
        if req and req.status in ('APROBADO', 'CONFIRMADO', 'VALIDADO'):
            # Preserve approved/frozen requisition so purchase forecast is never overwritten!
            return req

        if not req:
            req = Requisition(
                rotation_id=rotation.id,
                week=rotation.week,
                title=title or f"Requisición de Agroquímicos Semana {rotation.week}",
                status="PEDIDO_INICIAL",
                total_liters=round(total_liters_all, 1)
            )
            db.session.add(req)
            db.session.flush()
        else:
            req.total_liters = round(total_liters_all, 1)
            RequisitionItem.query.filter_by(requisition_id=req.id).delete()
            db.session.flush()

        for p_info in product_totals.values():
            qty_f = round_product_amount(p_info['quantity_forecast'], p_info['dose_unit'])
            item = RequisitionItem(
                requisition_id=req.id,
                product_id=p_info['product_id'],
                product_code=p_info['product_code'],
                commercial_name=p_info['commercial_name'],
                average_dose=p_info['dose'],
                unit=p_info['dose_unit'],
                quantity_forecast=qty_f,
                quantity_final=qty_f,
                difference=0.0,
                pest=p_info['pest']
            )
            db.session.add(item)

        db.session.commit()
        return req

    @classmethod
    def get_comparison_data(cls, rotation_id: int, cached_crop_records=None):
        """
        Builds the unified comparative data between:
        1. Cantidad Comprada (Requisición 15 días)
        2. Cantidad Ejecutada Real (Órdenes de Rotación generadas con el nuevo plano)
        3. Aplicaciones Adicionales / Extras de la semana
        4. Total Requerido Real = Ejecutado + Extras
        5. Diferencia = Total Requerido Real - Comprado
        """
        rotation = Rotation.query.get_or_404(rotation_id)
        req = Requisition.query.filter_by(rotation_id=rotation.id).first()
        if not req:
            req = cls.generate_or_update_forecast(rotation_id, cached_crop_records=cached_crop_records)

        # Forecast items from purchase order
        forecast_by_code = {it.product_code: it for it in req.items}

        # Real quantities from generated orders
        orders = FumigationOrder.query.filter_by(rotation_id=rotation.id).all()
        orders_by_code = {}
        for ord in orders:
            for ps in ord.products_summary:
                c = ps.product_code
                if c not in orders_by_code:
                    orders_by_code[c] = {
                        'product_code': c,
                        'commercial_name': ps.commercial_name,
                        'unit': ps.dose_unit,
                        'quantity_orders': 0.0,
                        'dose': ps.dose,
                        'pest': ps.pest
                    }
                orders_by_code[c]['quantity_orders'] += ps.total_required_quantity

        # Additional / Spot applications for this week
        add_apps = AdditionalApplication.query.filter_by(week=rotation.week).all()
        extras_by_code = {}
        for a in add_apps:
            c = a.product_code
            if c not in extras_by_code:
                extras_by_code[c] = {
                    'product_code': c,
                    'unit': a.dose_unit,
                    'quantity_extras': 0.0,
                    'reasons': []
                }
            extras_by_code[c]['quantity_extras'] += a.total_product
            if a.reason and a.reason not in extras_by_code[c]['reasons']:
                extras_by_code[c]['reasons'].append(a.reason)

        # Combine all product codes
        all_codes = list(set(list(forecast_by_code.keys()) + list(orders_by_code.keys()) + list(extras_by_code.keys())))
        all_codes.sort()

        # Batch load all products to avoid N+1 queries
        products_map = {p.code: p for p in Product.query.all()}

        comparison_rows = []
        for code in all_codes:
            f_item = forecast_by_code.get(code)
            o_item = orders_by_code.get(code)
            e_item = extras_by_code.get(code)

            # Determine product info
            prod = products_map.get(code)
            comm_name = prod.commercial_name if prod else (f_item.commercial_name if f_item else (o_item['commercial_name'] if o_item else code))
            unit = prod.unit if prod else (f_item.unit if f_item else (o_item['unit'] if o_item else (e_item['unit'] if e_item else 'CC')))
            pest = prod.pest if prod else (f_item.pest if f_item else (o_item['pest'] if o_item else 'Adicional'))

            qty_forecast = f_item.quantity_forecast if f_item else 0.0
            qty_orders = o_item['quantity_orders'] if o_item else 0.0
            qty_extras = e_item['quantity_extras'] if e_item else 0.0

            total_real = qty_orders + qty_extras

            # Determine product origin
            extra_reasons_list = e_item['reasons'] if e_item else []
            extra_reasons_str = ", ".join(extra_reasons_list) if extra_reasons_list else ""

            if qty_forecast > 0 and qty_extras > 0:
                origin = 'AMBOS'
                origin_label = 'Rotación + Extra'
                origin_badge = 'bg-dark text-white'
            elif qty_forecast > 0:
                origin = 'ROTACION'
                origin_label = 'Rotación (15 Días)'
                origin_badge = 'bg-primary text-white'
            elif qty_extras > 0:
                origin = 'EXTRA'
                origin_label = 'Extra / Mancha'
                origin_badge = 'bg-danger text-white'
            else:
                origin = 'OTRO'
                origin_label = 'Orden Directa'
                origin_badge = 'bg-secondary text-white'

            # Difference and Superavit / Deficit logic
            # If Comprado > Ejecutando -> Superávit (Sobra producto) -> Verde Fosforescente
            # If Ejecutando > Comprado -> Déficit (Falta comprar) -> Rojo
            if qty_forecast > total_real:
                diff_type = 'SUPERAVIT'
                diff_val = round_product_amount(qty_forecast - total_real, unit)
                diff_display = f"+{diff_val}"
                diff_badge_text = f"+{diff_val} (Sobra)"
                diff_cell_class = 'diff-cell-superavit'
                status_label = f'SUPERÁVIT (+{diff_val} {unit})'
            elif total_real > qty_forecast:
                diff_type = 'DEFICIT'
                diff_val = round_product_amount(total_real - qty_forecast, unit)
                diff_display = f"-{diff_val}"
                diff_badge_text = f"-{diff_val} (Falta)"
                diff_cell_class = 'diff-cell-deficit'
                status_label = f'DÉFICIT (-{diff_val} {unit})'
            else:
                diff_type = 'EXACTO'
                diff_val = 0.0
                diff_display = "0.0"
                diff_badge_text = "0.0 (Exacto)"
                diff_cell_class = 'diff-cell-neutral'
                status_label = 'EXACTO (0.0)'

            comparison_rows.append({
                'product_code': code,
                'commercial_name': comm_name,
                'unit': unit,
                'origin': origin,
                'origin_label': origin_label,
                'origin_badge': origin_badge,
                'extra_reasons_str': extra_reasons_str,
                'quantity_forecast': round_product_amount(qty_forecast, unit),
                'quantity_orders': round_product_amount(qty_orders, unit),
                'quantity_extras': round_product_amount(qty_extras, unit),
                'total_real': round_product_amount(total_real, unit),
                'difference': diff_val,
                'diff_type': diff_type,
                'diff_display': diff_display,
                'diff_badge_text': diff_badge_text,
                'diff_cell_class': diff_cell_class,
                'status': diff_type,
                'status_label': status_label,
                'pest': pest
            })

        # Calculate totals
        total_forecast = sum(r['quantity_forecast'] for r in comparison_rows)
        total_orders = sum(r['quantity_orders'] for r in comparison_rows)
        total_extras = sum(r['quantity_extras'] for r in comparison_rows)
        total_real_all = sum(r['total_real'] for r in comparison_rows)

        return {
            'requisition': req,
            'rotation': rotation,
            'rows': comparison_rows,
            'totals': {
                'total_forecast': round(total_forecast, 2),
                'total_orders': round(total_orders, 2),
                'total_extras': round(total_extras, 2),
                'total_real': round(total_real_all, 2)
            }
        }

    @classmethod
    def get_variety_breakdown_data(cls, rotation_id: int, cached_crop_records=None) -> list:
        """
        Builds a comprehensive breakdown of chemical allocation per crop / variety and phenological stage,
        enabling agronomists to quickly see how much chemical is assigned to each crop/variety (e.g. VERONICA, HYPERICUM).
        """
        rotation = Rotation.query.get_or_404(rotation_id)
        breakdown_dict = {}
        
        # Batch load products to avoid N+1 queries
        all_prods = Product.query.all()
        products_by_id = {p.id: p for p in all_prods}
        products_by_code = {p.code: p for p in all_prods}

        # 1. From rotation rounds
        for r in rotation.rounds:
            calc = CalculationEngine.calculate_round(r, cached_crop_records=cached_crop_records)
            for seg in calc.get('segments', []):
                c_name = (seg.get('crop_name') or 'SIN CULTIVO').strip().upper()
                stage = (seg.get('phenological_stage') or 'VEGETATIVO').strip().upper()
                std_beds = float(seg.get('standard_beds') or 0.0)
                tot_liters = float(seg.get('total_liters') or 0.0)
                
                for p in seg.get('products_detail', []):
                    p_code = p.get('product_code')
                    if not p_code:
                        continue
                    key = (c_name, stage, p_code)
                    if key not in breakdown_dict:
                        breakdown_dict[key] = {
                            'crop_name': c_name,
                            'stage': stage,
                            'product_code': p_code,
                            'commercial_name': p.get('commercial_name') or p_code,
                            'dose': p.get('dose', 0.0),
                            'unit': p.get('dose_unit') or 'CC',
                            'standard_beds': 0.0,
                            'total_liters': 0.0,
                            'total_product': 0.0,
                            'rounds_list': []
                        }
                    breakdown_dict[key]['standard_beds'] += std_beds
                    breakdown_dict[key]['total_liters'] += tot_liters
                    breakdown_dict[key]['total_product'] += float(p.get('product_amount') or 0.0)
                    r_label = f"{r.name} ({r.scheduled_day})"
                    if r_label not in breakdown_dict[key]['rounds_list']:
                        breakdown_dict[key]['rounds_list'].append(r_label)

        # 2. From extra applications
        add_apps = AdditionalApplication.query.filter_by(week=rotation.week).all()
        for a in add_apps:
            c_name = (a.crop_name or 'EXTRA / GENERAL').strip().upper()
            stage = 'EXTRA'
            p_code = a.product_code
            if not p_code:
                continue
            key = (c_name, stage, p_code)
            if key not in breakdown_dict:
                prod = products_by_id.get(a.product_id) if a.product_id else products_by_code.get(p_code)
                comm_name = prod.commercial_name if prod else p_code
                unit = prod.unit if prod else (a.dose_unit or 'CC')
                breakdown_dict[key] = {
                    'crop_name': c_name,
                    'stage': stage,
                    'product_code': p_code,
                    'commercial_name': comm_name,
                    'dose': a.dose_applied,
                    'unit': unit,
                    'standard_beds': 0.0,
                    'total_liters': 0.0,
                    'total_product': 0.0,
                    'rounds_list': ['Aplicación Extra / Mancha']
                }
            breakdown_dict[key]['standard_beds'] += a.standard_beds
            breakdown_dict[key]['total_liters'] += a.total_liters
            breakdown_dict[key]['total_product'] += a.total_product
            if 'Aplicación Extra / Mancha' not in breakdown_dict[key]['rounds_list']:
                breakdown_dict[key]['rounds_list'].append('Aplicación Extra / Mancha')

        results = []
        for key, item in breakdown_dict.items():
            u = item['unit']
            results.append({
                'crop_name': item['crop_name'],
                'stage': item['stage'],
                'product_code': item['product_code'],
                'commercial_name': item['commercial_name'],
                'dose': item['dose'],
                'unit': u,
                'standard_beds': round(item['standard_beds'], 2),
                'total_liters': round(item['total_liters'], 1),
                'total_product': round_product_amount(item['total_product'], u),
                'rounds_str': ", ".join(item['rounds_list'])
            })
            
        results.sort(key=lambda x: (x['crop_name'], x['stage'], x['product_code']))
        return results

    @classmethod
    def get_consolidated_pivot_data(cls, week: str) -> dict:
        """
        Builds a multi-module dynamic pivot table for a given week:
        Consolidates Demand across:
        - 15-Day Purchase Order (Requisition forecast)
        - Fumigación Foliar (Weekly Rotation Orders + Spot Extra Applications)
        - Drench (Soil / Root applications)
        - Trichos (Biological inoculations)
        
        Outputs consolidated totals and itemized breakdowns per module and crop.
        """
        week = (week or '').strip()
        rotation = Rotation.query.filter_by(week=week).order_by(Rotation.version.desc()).first()
        
        req = None
        forecast_by_code = {}
        if rotation:
            req = Requisition.query.filter_by(rotation_id=rotation.id).first()
            if not req:
                req = cls.generate_or_update_forecast(rotation.id)
            if req:
                forecast_by_code = {it.product_code: it for it in req.items}

        # 1. Fumigación Orders
        fumi_by_code = {}
        fumi_details_by_code = {}
        if rotation:
            orders = FumigationOrder.query.filter_by(rotation_id=rotation.id).all()
            for ord in orders:
                for ps in ord.products_summary:
                    c = ps.product_code
                    if not c:
                        continue
                    if c not in fumi_by_code:
                        fumi_by_code[c] = 0.0
                        fumi_details_by_code[c] = []
                    fumi_by_code[c] += ps.total_required_quantity
                    fumi_details_by_code[c].append({
                        'module': 'Fumigación Foliar',
                        'program': f"Orden {ord.order_number} (Vuelta {ord.round_number})",
                        'crop_name': 'Rotación General',
                        'beds': 0.0,
                        'liters': 0.0,
                        'amount': ps.total_required_quantity,
                        'unit': ps.dose_unit
                    })

        # 2. Extras / Manchas
        add_apps = AdditionalApplication.query.filter_by(week=week).all()
        for a in add_apps:
            c = a.product_code
            if not c:
                continue
            if c not in fumi_by_code:
                fumi_by_code[c] = 0.0
                fumi_details_by_code[c] = []
            fumi_by_code[c] += a.total_product
            fumi_details_by_code[c].append({
                'module': 'Mancha / Extra',
                'program': f"Mancha Bloque {a.block_name} ({a.reason or 'Adicional'})",
                'crop_name': a.crop_name or 'Extra',
                'beds': a.standard_beds,
                'liters': a.total_liters,
                'amount': a.total_product,
                'unit': a.dose_unit or 'CC'
            })

        # 3. Drench (Prepared for Drench module integration)
        drench_by_code = {}
        drench_details_by_code = {}

        # 4. Trichos / Biológicos (Prepared for Trichos module integration)
        trichos_by_code = {}
        trichos_details_by_code = {}

        # Combine all codes
        all_codes = list(set(
            list(forecast_by_code.keys()) + 
            list(fumi_by_code.keys()) + 
            list(drench_by_code.keys()) + 
            list(trichos_by_code.keys())
        ))
        all_codes.sort()

        pivot_rows = []
        for code in all_codes:
            f_item = forecast_by_code.get(code)
            prod = Product.query.filter_by(code=code).first()
            comm_name = prod.commercial_name if prod else (f_item.commercial_name if f_item else code)
            unit = prod.unit if prod else (f_item.unit if f_item else 'CC')

            qty_forecast = f_item.quantity_forecast if f_item else 0.0
            qty_fumi = fumi_by_code.get(code, 0.0)
            qty_drench = drench_by_code.get(code, 0.0)
            qty_trichos = trichos_by_code.get(code, 0.0)

            total_real = qty_fumi + qty_drench + qty_trichos

            if qty_forecast > total_real:
                diff_type = 'SUPERAVIT'
                diff_val = round_product_amount(qty_forecast - total_real, unit)
                diff_display = f"+{diff_val}"
                diff_badge_text = f"+{diff_val} (Sobra)"
                diff_cell_class = 'diff-cell-superavit'
                status_label = f'SUPERÁVIT (+{diff_val} {unit})'
            elif total_real > qty_forecast:
                diff_type = 'DEFICIT'
                diff_val = round_product_amount(total_real - qty_forecast, unit)
                diff_display = f"-{diff_val}"
                diff_badge_text = f"-{diff_val} (Falta)"
                diff_cell_class = 'diff-cell-deficit'
                status_label = f'DÉFICIT (-{diff_val} {unit})'
            else:
                diff_type = 'EXACTO'
                diff_val = 0.0
                diff_display = "0.0"
                diff_badge_text = "0.0 (Exacto)"
                diff_cell_class = 'diff-cell-neutral'
                status_label = 'EXACTO (0.0)'

            # Details list for dynamic child expansion
            details_list = (
                fumi_details_by_code.get(code, []) +
                drench_details_by_code.get(code, []) +
                trichos_details_by_code.get(code, [])
            )

            pivot_rows.append({
                'product_code': code,
                'commercial_name': comm_name,
                'unit': unit,
                'quantity_forecast': round_product_amount(qty_forecast, unit),
                'quantity_fumigacion': round_product_amount(qty_fumi, unit),
                'quantity_drench': round_product_amount(qty_drench, unit),
                'quantity_trichos': round_product_amount(qty_trichos, unit),
                'total_real': round_product_amount(total_real, unit),
                'difference': diff_val,
                'diff_type': diff_type,
                'diff_display': diff_display,
                'diff_badge_text': diff_badge_text,
                'diff_cell_class': diff_cell_class,
                'status_label': status_label,
                'details': details_list
            })

        total_forecast = sum(r['quantity_forecast'] for r in pivot_rows)
        total_fumi = sum(r['quantity_fumigacion'] for r in pivot_rows)
        total_drench = sum(r['quantity_drench'] for r in pivot_rows)
        total_trichos = sum(r['quantity_trichos'] for r in pivot_rows)
        total_real_all = sum(r['total_real'] for r in pivot_rows)

        return {
            'week': week,
            'rotation': rotation,
            'requisition': req,
            'rows': pivot_rows,
            'totals': {
                'total_forecast': round(total_forecast, 2),
                'total_fumigacion': round(total_fumi, 2),
                'total_drench': round(total_drench, 2),
                'total_trichos': round(total_trichos, 2),
                'total_real': round(total_real_all, 2)
            }
        }

    @classmethod
    def sync_with_final_orders(cls, rotation_id: int):
        """
        Synchronizes the Requisition model with the latest state of orders and extras.
        """
        data = cls.get_comparison_data(rotation_id)
        req = data['requisition']

        for row in data['rows']:
            it = RequisitionItem.query.filter_by(requisition_id=req.id, product_code=row['product_code']).first()
            if it:
                it.quantity_final = row['total_real']
                it.difference = row['difference']
            else:
                prod = Product.query.filter_by(code=row['product_code']).first()
                new_it = RequisitionItem(
                    requisition_id=req.id,
                    product_id=prod.id if prod else None,
                    product_code=row['product_code'],
                    commercial_name=row['commercial_name'],
                    average_dose=prod.dose_fumigation if prod else 0.0,
                    unit=row['unit'],
                    quantity_forecast=row['quantity_forecast'],
                    quantity_final=row['total_real'],
                    difference=row['difference'],
                    pest=row['pest']
                )
                db.session.add(new_it)

        req.status = "VALIDADO"
        db.session.commit()
        return req

    @staticmethod
    def export_requisition_to_excel(req_obj) -> io.BytesIO:
        """
        Exports the purchase comparison and crop chemical allocation to an Excel workbook (.xlsx).
        Sheet 1: Comparativa_Requisicion (Producto, Nombre Comercial, Origen, Comprado, Ejecutando, Diferencia, Unidad)
        Sheet 2: Asignacion_Por_Cultivo (Desglose de químicos por cultivo y variedad)
        """
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        comp_data = RequisitionService.get_comparison_data(req_obj.rotation_id)
        variety_data = RequisitionService.get_variety_breakdown_data(req_obj.rotation_id)

        rows_comp = []
        for r in comp_data['rows']:
            u = r['unit']
            is_int = is_integer_unit(u)
            rows_comp.append({
                'PRODUCTO QUÍMICO': r['product_code'],
                'NOMBRE COMERCIAL': r['commercial_name'],
                'UNIDAD': r['unit'],
                'ORIGEN': r['origin_label'],
                'COMPRADO (15 DÍAS)': int(round(r['quantity_forecast'])) if is_int else round(r['quantity_forecast'], 2),
                'EJECUTANDO (REAL)': int(round(r['total_real'])) if is_int else round(r['total_real'], 2),
                'DIFERENCIA': r['diff_display']
            })

        rows_var = []
        for v in variety_data:
            u = v['unit']
            is_int = is_integer_unit(u)
            rows_var.append({
                'CULTIVO / VARIEDAD': v['crop_name'],
                'ESTADO FENOLÓGICO': v['stage'],
                'CÓDIGO PRODUCTO': v['product_code'],
                'NOMBRE COMERCIAL': v['commercial_name'],
                'UNIDAD': v['unit'],
                'CAMAS STD': v['standard_beds'],
                'LITROS MEZCLA': v['total_liters'],
                'CANTIDAD ASIGNADA': int(round(v['total_product'])) if is_int else round(v['total_product'], 2)
            })

        df_comp = pd.DataFrame(rows_comp)
        df_var = pd.DataFrame(rows_var)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            sheet_comp_name = f"Comparativa_{req_obj.week}"
            sheet_var_name = "Asignacion_Por_Cultivo"

            df_comp.to_excel(writer, sheet_name=sheet_comp_name, index=False)
            df_var.to_excel(writer, sheet_name=sheet_var_name, index=False)

            ws_comp = writer.sheets[sheet_comp_name]
            ws_var = writer.sheets[sheet_var_name]

            # Style Sheet 1: Comparativa
            fill_superavit = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
            fill_deficit = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
            font_white = Font(color="FFFFFF", bold=True)
            font_header = Font(color="FFFFFF", bold=True)
            fill_header = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")

            for col_idx in range(1, len(df_comp.columns) + 1):
                cell = ws_comp.cell(row=1, column=col_idx)
                cell.fill = fill_header
                cell.font = font_header
                cell.alignment = Alignment(horizontal="center")

            # Color difference column (Column 7 = DIFERENCIA)
            for row_idx, r in enumerate(comp_data['rows'], start=2):
                diff_cell = ws_comp.cell(row=row_idx, column=7)  # Diferencia column
                if r['diff_type'] == 'SUPERAVIT':
                    diff_cell.fill = fill_superavit
                    diff_cell.font = font_white
                elif r['diff_type'] == 'DEFICIT':
                    diff_cell.fill = fill_deficit
                    diff_cell.font = font_white
                    diff_cell.font = font_white

            for col in ws_comp.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = col[0].column_letter
                ws_comp.column_dimensions[col_letter].width = max(max_len + 3, 14)

            # Style Sheet 2: Variety breakdown
            for col_idx in range(1, len(df_var.columns) + 1):
                cell = ws_var.cell(row=1, column=col_idx)
                cell.fill = fill_header
                cell.font = font_header
                cell.alignment = Alignment(horizontal="center")

            for col in ws_var.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = col[0].column_letter
                ws_var.column_dimensions[col_letter].width = max(max_len + 3, 14)

        output.seek(0)
        return output

    @classmethod
    def recalculate_with_budget(cls, requisition_id: int, budget_adjustments: list):
        """
        Recalcula la requisición sumando las camas de presupuesto.
        
        Args:
            requisition_id: ID de la requisición
            budget_adjustments: Lista de ajustes [{crop_name, age, stage, beds, reason}, ...]
        """
        requisition = Requisition.query.get_or_404(requisition_id)
        rotation = Rotation.query.get_or_404(requisition.rotation_id)
        
        # First, regenerate base requisition from rotation (without budget)
        product_totals = {}
        total_liters_all = 0.0
        
        # Calculate from rotation rounds
        for r in rotation.rounds:
            calc = CalculationEngine.calculate_round(r)
            total_liters_all += calc['totals']['total_liters']
            
            for ps in calc['product_summaries']:
                key = ps['product_code']
                if key not in product_totals:
                    product_totals[key] = {
                        'product_id': ps['product_id'],
                        'product_code': ps['product_code'],
                        'commercial_name': ps['commercial_name'],
                        'dose': ps['dose'],
                        'dose_unit': ps['dose_unit'],
                        'quantity_forecast': 0.0,
                        'pest': ps['pest']
                    }
                product_totals[key]['quantity_forecast'] += ps['total_required_quantity']
        
        # Now add budget adjustments
        # For each adjustment, calculate additional product needs
        for adj in budget_adjustments:
            crop_name = adj.get('crop_name')
            age = adj.get('age')
            stage = adj.get('stage')
            beds = adj.get('beds', 0)
            
            if not all([crop_name, age, stage, beds]):
                continue
            
            # Get liters per bed for this crop/age
            from app.shared.models import Litraje
            lit_rule = Litraje.query.filter_by(crop_name=crop_name, age=int(age)).first()
            liters_per_bed = lit_rule.liters_per_bed if lit_rule else 80.0  # Default 80L
            
            total_liters_budget = beds * liters_per_bed
            total_liters_all += total_liters_budget
            
            # Get products for this crop/stage from rotation rounds
            # We'll use the products from the first round that has this crop/stage
            products_to_apply = []
            for r in rotation.rounds:
                for item in r.items:
                    if item.crop_name == crop_name and item.phenological_stage == stage:
                        products_to_apply.append({
                            'product_id': item.product_id,
                            'product_code': item.product.code,
                            'commercial_name': item.product.commercial_name,
                            'dose': item.dose_applied,
                            'dose_unit': item.dose_unit,
                            'pest': item.product.pest
                        })
                if products_to_apply:
                    break  # Use products from first matching round
            
            # Calculate additional quantities for these products
            for prod_info in products_to_apply:
                key = prod_info['product_code']
                dose = prod_info['dose']
                
                # Calculate additional quantity needed
                additional_qty = (dose * total_liters_budget) / 100.0
                
                if key not in product_totals:
                    product_totals[key] = {
                        'product_id': prod_info['product_id'],
                        'product_code': prod_info['product_code'],
                        'commercial_name': prod_info['commercial_name'],
                        'dose': dose,
                        'dose_unit': prod_info['dose_unit'],
                        'quantity_forecast': 0.0,
                        'pest': prod_info['pest']
                    }
                
                product_totals[key]['quantity_forecast'] += additional_qty
        
        # Update requisition
        requisition.total_liters = round(total_liters_all, 1)
        
        # Delete old items and create new ones with updated quantities
        RequisitionItem.query.filter_by(requisition_id=requisition.id).delete()
        db.session.flush()
        
        for p_info in product_totals.values():
            item = RequisitionItem(
                requisition_id=requisition.id,
                product_id=p_info['product_id'],
                product_code=p_info['product_code'],
                commercial_name=p_info['commercial_name'],
                average_dose=p_info['dose'],
                unit=p_info['dose_unit'],
                quantity_forecast=round(p_info['quantity_forecast'], 2),
                quantity_final=round(p_info['quantity_forecast'], 2),
                difference=0.0,
                pest=p_info['pest']
            )
            db.session.add(item)
        
        db.session.commit()
        return requisition
